import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from config import PATH_REPORTES

def generar_excel(df, factura, nit):
    if df is None or df.empty:
        return

    nombre_archivo = f"CRUCE_{factura}_{nit}.xlsx"
    ruta_completa = os.path.join(PATH_REPORTES, nombre_archivo)

    # --------------------------------------------------------
    # 1. PREPARACIÓN DE DATOS MATEMÁTICOS
    # --------------------------------------------------------
    # Excluimos la fila de totales globales del final si existe
    df_items = df[df['codigo_barras'] != 'TOTAL FACTURA'].copy()
    
    # --- PESTAÑA NP (Novedad Producto/Cantidad) ---
    df_np = df_items.copy()
    df_np['COSTO_UNT'] = df_np['erp_precio'].where(df_np['erp_precio'] > 0, df_np['xml_precio'])
    df_np['DIF_UND'] = df_np['xml_cant'] - df_np['erp_cant']
    df_np['COSTO_TOTAL'] = df_np['DIF_UND'] * df_np['COSTO_UNT']

    # --- PESTAÑA AC (Análisis Costo) ---
    df_ac = df_items.copy()
    
    # El ajuste de costo es la diferencia TOTAL de dinero, menos lo que ya se explicó por cantidades en NP
    df_ac['COSTO_DIF_TOTAL'] = df_ac['dif_total'] - df_np['COSTO_TOTAL']
    
    # Calculamos la diferencia unitaria equivalente (y evitamos dividir por cero)
    df_ac['DIF_COSTO_UND'] = df_ac['COSTO_DIF_TOTAL'] / df_ac['xml_cant'].replace(0, 1) 
    
    # --- APLICAMOS EL FORMATO Y FILTROS A LAS TABLAS ---
    
    # Columnas de AC
    df_ac = df_ac[[
        'item_xml', 'item_erp', 'descripcion_xml', 'descripcion_erp',
        'DIF_COSTO_UND', 'xml_cant', 'COSTO_DIF_TOTAL'
    ]]
    df_ac.columns = [
        'ITEM_XML', 'ITEM_ERP', 'DESCRIPCION_XML', 'DESCRIPCION_ERP',
        'DIF_COSTO_UND', 'CANTIDAD_XML', 'COSTO_DIF_TOTAL'
    ]
    # FILTRO: Mandamos a la pestaña AC cualquier descuadre de dinero mayor o igual a 1 peso
    df_ac = df_ac[abs(df_ac['COSTO_DIF_TOTAL']) >= 1]

    # Columnas de NP
    df_np = df_np[[
        'item_xml', 'item_erp', 'descripcion_xml', 'descripcion_erp',
        'COSTO_UNT', 'erp_cant', 'xml_cant', 'DIF_UND', 'COSTO_TOTAL'
    ]]
    df_np.columns = [
        'ITEM_XML', 'ITEM_ERP', 'DESCRIPCION_XML', 'DESCRIPCION_ERP',
        'COSTO_UNT', 'CANTIDAD_ERP', 'CANTIDAD_XML', 'DIF_UND', 'COSTO_TOTAL'
    ]
    # FILTRO: Solo pasamos a NP si hubo diferencia de unidades físicas
    df_np = df_np[abs(df_np['DIF_UND']) > 0]

    # --- CÁLCULOS PARA LA PORTADA ---
    # Valores Brutos Globales
    valor_xml = df_items['xml_total'].sum()
    valor_erp = df_items['erp_total'].sum()
    dif_valor = valor_xml - valor_erp
    
    cant_xml = df_items['xml_cant'].sum()
    cant_erp = df_items['erp_cant'].sum()
    dif_cant = cant_xml - cant_erp
    
    # Ajustes Acumulados
    ajuste_costo_ac = df_ac['COSTO_DIF_TOTAL'].sum() if not df_ac.empty else 0
    ajuste_costo_np = df_np['COSTO_TOTAL'].sum() if not df_np.empty else 0
    total_ajuste_costo = ajuste_costo_ac + ajuste_costo_np
    saldo_dif_costo = dif_valor - total_ajuste_costo

    ajuste_cant_ac = 0 # La pestaña AC es solo costos, no mueve cantidades
    ajuste_cant_np = df_np['DIF_UND'].sum() if not df_np.empty else 0
    total_ajuste_cant = ajuste_cant_ac + ajuste_cant_np
    saldo_dif_cant = dif_cant - total_ajuste_cant

    # --------------------------------------------------------
    # 2. ESCRITURA BÁSICA CON PANDAS
    # --------------------------------------------------------
    with pd.ExcelWriter(ruta_completa, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Detalle")
        df_ac.to_excel(writer, index=False, sheet_name="AC")
        df_np.to_excel(writer, index=False, sheet_name="NP")

    # --------------------------------------------------------
    # 3. FORMATO VISUAL Y CREACIÓN DE PORTADA (OPENPYXL)
    # --------------------------------------------------------
    wb = load_workbook(ruta_completa)
    
    # Definición de Estilos Globales
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 
    rojo_fuerte = PatternFill(start_color="C00000", fill_type="solid") 
    verde_fuerte = PatternFill(start_color="008000", fill_type="solid") 
    blanco_font = Font(bold=True, size=12, color="FFFFFF")
    titulo_gigante = Font(bold=True, size=36)
    borde_fino = Side(border_style="thin", color="000000")
    borde = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
    formato_moneda = '"$"#,##0.00'

    # --- 3.1 CREACIÓN DE LA PORTADA ---
    ws_portada = wb.create_sheet("PORTADA", 0) # La crea en la posición 0 (Primera)
    wb.active = ws_portada # Que el archivo abra en esta pestaña

    # Anchos de columna para la portada
    ws_portada.column_dimensions['A'].width = 3
    ws_portada.column_dimensions['B'].width = 35
    ws_portada.column_dimensions['C'].width = 25
    ws_portada.column_dimensions['D'].width = 25
    ws_portada.column_dimensions['E'].width = 5 # Espacio divisor
    ws_portada.column_dimensions['F'].width = 35
    ws_portada.column_dimensions['G'].width = 20
    ws_portada.column_dimensions['H'].width = 25

    # Título Principal (Dinámico según saldo)
    titulo_portada = ws_portada['C2']
    if abs(dif_valor) > 1 or abs(dif_cant) > 0:
        titulo_portada.value = "NECESARIO VALIDACION MANUAL"
        titulo_portada.fill = amarillo
        titulo_portada.font = Font(bold=True, color="000000")
    else:
        titulo_portada.value = "CUADRE PERFECTO - SIN ACCIÓN REQUERIDA"
        titulo_portada.fill = verde_fuerte
        titulo_portada.font = blanco_font
    ws_portada.merge_cells('C2:G2')
    titulo_portada.alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulos
    ws_portada['C4'].value = "COSTO"
    ws_portada['C4'].alignment = Alignment(horizontal='center')
    ws_portada['G4'].value = "UNIDADES"
    ws_portada['G4'].alignment = Alignment(horizontal='center')

    ws_portada['B6'].value = "PORTADA"
    ws_portada['H6'].value = "REALIDAD DE LA FACTURA"
    ws_portada['H6'].alignment = Alignment(horizontal='right')

    # SECCIÓN: VALORES BRUTOS
    ws_portada['B8'].value = "VALOR_XML"
    ws_portada['C8'].value = valor_xml
    ws_portada['C8'].number_format = formato_moneda
    ws_portada['B9'].value = "VALOR_ERP"
    ws_portada['C9'].value = valor_erp
    ws_portada['C9'].number_format = formato_moneda

    ws_portada['F8'].value = "CANTIDAD_XML"
    ws_portada['G8'].value = cant_xml
    ws_portada['F9'].value = "CANTIDAD_ERP"
    ws_portada['G9'].value = cant_erp

    # SECCIÓN: DIFERENCIAS (Con número gigante)
    ws_portada['B11'].value = "DIFERENCIA"
    ws_portada['B11'].alignment = Alignment(vertical='center')
    ws_portada['C11'].value = dif_valor
    ws_portada['C11'].number_format = formato_moneda
    ws_portada['C11'].font = titulo_gigante
    ws_portada['C11'].alignment = Alignment(horizontal='center')
    ws_portada.row_dimensions[11].height = 45

    ws_portada['F11'].value = "DIFERENCIA"
    ws_portada['F11'].alignment = Alignment(vertical='center')
    ws_portada['G11'].value = dif_cant
    ws_portada['G11'].font = titulo_gigante
    ws_portada['G11'].alignment = Alignment(horizontal='center')

    # Mensajes "SIN ACCIONES SUGERIDAS"
    if abs(dif_valor) <= 1: ws_portada['D11'].value = "SIN ACCIONES SUGERIDAS"
    if dif_cant == 0: ws_portada['H11'].value = "SIN ACCIONES SUGERIDAS"

    # SECCIÓN: AJUSTES SUGERIDOS
    ws_portada['B13'].value = "AJUSTE COSTO SUGERIDO AC"
    ws_portada['C13'].value = ajuste_costo_ac
    ws_portada['C13'].number_format = formato_moneda
    ws_portada['B14'].value = "AJUSTE NO PEDIDO / DEVOLUCIONES"
    ws_portada['C14'].value = ajuste_costo_np
    ws_portada['C14'].number_format = formato_moneda
    ws_portada['B15'].value = "TOTAL AJUSTE SUGERIDO"
    ws_portada['C15'].value = total_ajuste_costo
    ws_portada['C15'].number_format = formato_moneda

    ws_portada['F13'].value = "AJUSTE CANTIDAD SUGERIDA AC"
    ws_portada['G13'].value = ajuste_cant_ac
    ws_portada['F14'].value = "AJUSTE NO PEDIDO / DEVOLUCIONES"
    ws_portada['G14'].value = ajuste_cant_np
    ws_portada['F15'].value = "TOTAL AJUSTE SUGERIDO"
    ws_portada['G15'].value = total_ajuste_cant
    ws_portada['H13'].value = "ACCIONES SUGERIDAS"

    # SECCIÓN: SALDOS FINALES
    ws_portada['B18'].value = "SALDO EN DIFERENCIA"
    ws_portada['C18'].value = saldo_dif_costo
    ws_portada['C18'].number_format = formato_moneda
    ws_portada['F18'].value = "SALDO EN DIFERENCIA"
    ws_portada['G18'].value = saldo_dif_cant

    # Aplicar Bordes Decorativos de la Portada
    cajas = [('B6','D11'), ('F6','H11'), ('B13','D15'), ('F13','H15'), ('B18','D18'), ('F18','H18')]
    for inicio, fin in cajas:
        rango = ws_portada[f'{inicio}:{fin}']
        for row in rango:
            for cell in row:
                top = borde_fino if cell.row == ws_portada[inicio].row else None
                bottom = borde_fino if cell.row == ws_portada[fin].row else None
                left = borde_fino if cell.column == ws_portada[inicio].column else None
                right = borde_fino if cell.column == ws_portada[fin].column else None
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # --- 3.2 FORMATO HOJA DETALLE ---
    if "Detalle" in wb.sheetnames:
        ws = wb["Detalle"]
        hay_errores = not df_items[df_items['estado'] != 'OK'].empty

        col_estado, col_alerta, col_codigo, col_item_erp = None, None, None, None
        for cell in ws[1]:
            if cell.value == "estado": col_estado = cell.column
            if cell.value == "alerta_cruce": col_alerta = cell.column
            if cell.value == "codigo_barras": col_codigo = cell.column
            if cell.value == "item_erp": col_item_erp = cell.column
        
        for row in range(2, ws.max_row + 1):
            c_estado = ws.cell(row, col_estado) if col_estado else None
            c_alerta = ws.cell(row, col_alerta) if col_alerta else None
            c_codigo = ws.cell(row, col_codigo) if col_codigo else None
            c_item_erp = ws.cell(row, col_item_erp) if col_item_erp else None
            
            # Fila "TOTAL FACTURA"
            if c_codigo and c_codigo.value == "TOTAL FACTURA":
                color_total = verde_fuerte if c_estado and c_estado.value == "OK" else rojo_fuerte
                for col in range(1, ws.max_column + 1):
                    cel = ws.cell(row, col)
                    cel.fill = color_total
                    cel.font = blanco_font
                    cel.border = borde
                continue 
            
            # Formato condicional celdas normales
            if c_estado:
                c_estado.fill = verde if c_estado.value == "OK" else rojo
            if c_alerta and c_alerta.value == "RESCATE CEROS":
                c_alerta.fill = amarillo
                if c_item_erp: c_item_erp.fill = amarillo
            
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).border = borde

        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column) 
        titulo = ws['A1']
        titulo.font = blanco_font
        titulo.alignment = Alignment(horizontal="center", vertical="center")
        titulo.value = "RESULTADO: DIFERENCIAS ENCONTRADAS" if hay_errores else "CRUCE EXITOSO - FACTURA CUADRADA"
        titulo.fill = rojo_fuerte if hay_errores else verde_fuerte
        ws.row_dimensions[1].height = 30

    # --- 3.3 FORMATO HOJAS AC Y NP ---
    for nombre_hoja in ["AC", "NP"]:
        if nombre_hoja in wb.sheetnames:
            ws_tmp = wb[nombre_hoja]
            for row in ws_tmp.iter_rows(min_row=2):
                for cell in row:
                    cell.border = borde

    # Ajustar ancho de columnas para las otras hojas
    for sheet in wb.sheetnames:
        if sheet == "PORTADA": continue
        ws_tmp = wb[sheet]
        for column_cells in ws_tmp.columns:
            try:
                length = max(len(str(cell.value) or "") for cell in column_cells)
                ws_tmp.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)
            except: pass

    wb.save(ruta_completa)
    print(f"Excel generado con Portada Dinámica y Ajustes AC/NP: {ruta_completa}")